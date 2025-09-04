import re
import csv
import sqlite3
import tkinter as tk
from tkinter import scrolledtext, filedialog, messagebox, ttk
import openpyxl
from openpyxl import Workbook
from datetime import datetime
import qrcode
from PIL import Image, ImageTk, ImageFont
import os
import sys
import webbrowser
import tkinter.font as tkfont

# ---------- Settings ----------
PERSIAN_TO_EN = str.maketrans('۰۱۲۳۴۵۶۷۸۹', '0123456789')
pattern = re.compile(r'(?:\+?98|0)?9[\d\-\s\(\)]{8,14}\d')

BG_COLOR = "#1e1e2f"
BTN_COLOR = "#3a3a5c"
BTN_HOVER = "#50507a"
TEXT_BG = "#2a2a40"
TEXT_FG = "#f1f1f1"

DB_FILE = "history.db"

# ---------- Language / i18n ----------
LANG_FA = True  # start Persian
T = {
    "title": ("Numify", "نامیفای"),
    "menu_file": ("File", "فایل"),
    "menu_exit": ("Exit", "خروج"),
    "menu_history": ("History", "سوابق"),
    "menu_history_show": ("Show History", "نمایش سوابق"),
    "menu_lang": ("Language", "زبان"),
    "menu_lang_toggle": ("English / فارسی", "فارسی / English"),
    "label_input": ("📥 Input Text:", "📥 متن ورودی:"),
    "label_prefix": ("🔤 Name Prefix:", "🔤 پیشوند نام مخاطبین:"),
    "group_input": ("📥 Input", "📥 ورودی"),
    "group_output": ("📤 Output", "📤 خروجی"),
    "btn_paste": ("📋 Paste", "📋 چسباندن"),
    "btn_extract_text": ("🔍 Extract from Text", "🔍 استخراج از متن"),
    "btn_load_excel": ("📂 Load Excel", "📂 بارگذاری Excel"),
    "btn_load_csv": ("📂 Load CSV", "📂 بارگذاری CSV"),
    "btn_save_vcf": ("💾 Save VCF", "💾 ذخیره VCF"),
    "btn_save_csv": ("📊 Save CSV", "📊 ذخیره CSV"),
    "btn_save_xlsx": ("📘 Save Excel", "📘 ذخیره Excel"),
    "btn_copy": ("📌 Copy", "📌 کپی"),
    "btn_history": ("📜 History", "📜 تاریخچه"),
    "btn_lang": ("🌐 Language", "🌐 زبان"),
    "btn_clear": ("🧹 Clear", "🧹 پاک‌سازی"),
    "btn_donate": ("💖 Donate", "💖 حمایت مالی"),
    "label_preview": ("📤 Extracted Numbers Preview:", "📤 پیش‌نمایش شماره‌های استخراج‌شده:"),
    "status_ready": ("📌 Ready...", "📌 آماده..."),
    "warn_no_numbers": ("⚠️ No numbers available.", "⚠️ هیچ شماره‌ای موجود نیست."),
    "saved_vcf": ("💾 VCF saved to:\n{}", "💾 VCF ذخیره شد:\n{}"),
    "saved_csv": ("💾 CSV saved to:\n{}", "💾 CSV ذخیره شد:\n{}"),
    "saved_xlsx": ("💾 Excel saved to:\n{}", "💾 Excel ذخیره شد:\n{}"),
    "copied": ("📌 Copied to clipboard.", "📌 در کلیپ‌بورد کپی شد."),
    "clipboard_empty": ("⚠️ Clipboard is empty.", "⚠️ کلیپ‌بورد خالی است."),
    "loaded_excel": ("✅ Numbers loaded from Excel:\n{}", "✅ شماره‌ها از اکسل بارگذاری شد:\n{}"),
    "loaded_csv": ("✅ Numbers loaded from CSV:\n{}", "✅ شماره‌ها از CSV بارگذاری شد:\n{}"),
    "error": ("❌ Error", "❌ خطا"),
    "history_window_title": ("📜 History", "📜 سوابق"),
    "history_dates": ("📅 Dates", "📅 تاریخ‌ها"),
    "history_id": ("ID", "ID"),
    "history_datetime": ("Date & Time", "تاریخ و زمان"),
    "history_prefix": ("🔤 Output Prefix:", "🔤 پیشوند خروجی:"),
    "history_count": ("📌 {} numbers", "📌 {} شماره"),
    "history_select_first": ("⚠️ Select a record first.", "⚠️ ابتدا یک تاریخ را انتخاب کنید."),
    "donate_title": ("💖 Donate", "💖 حمایت مالی"),
    "donate_thanks": ("🙏 Thank you for your support!", "🙏 از حمایت شما سپاسگزاریم!"),
    "donate_hint": ("You can donate via wallets or coffee link.", "می‌توانید از طریق ولت‌ها یا لینک کافیته حمایت کنید."),
    "link_open": ("🔗 Open link", "🔗 باز کردن لینک"),
    "close": ("❌ Close", "❌ بستن"),
}

def tr(key, *fmt):
    pair = T.get(key, ("", ""))
    s = pair[1] if LANG_FA else pair[0]
    return s.format(*fmt) if fmt else s

def toggle_language():
    global LANG_FA
    LANG_FA = not LANG_FA
    refresh_ui_texts()

# ---------- Font: load YekanBakh from relative path ----------
FONT_PATH = os.path.join("Numify", "font", "YekanBakh-Regular.ttf")
UI_FAMILY = None
def try_register_font_windows(path):
    try:
        import ctypes
        FR_PRIVATE = 0x10
        added = ctypes.windll.gdi32.AddFontResourceExW(os.path.abspath(path), FR_PRIVATE, 0)
        return added != 0
    except Exception:
        return False

def load_ui_font(root: tk.Tk):
    global UI_FAMILY
    # Try get family name from TTF
    family_guess = None
    try:
        pf = ImageFont.truetype(FONT_PATH, 12)
        family_guess = pf.getname()[0]
    except Exception:
        pass

    # Windows: attempt to register TTF privately
    if sys.platform.startswith("win") and os.path.isfile(FONT_PATH):
        try_register_font_windows(FONT_PATH)

    # If family exists in Tk, use it
    fams = set(tkfont.families(root))
    if family_guess and family_guess in fams:
        UI_FAMILY = family_guess
    else:
        # Some Yekan variants / fallback family names
        for name in ["Yekan Bakh", "YekanBakh", "IRANYekan", "Vazirmatn", "Vazir", "Segoe UI"]:
            if name in fams:
                UI_FAMILY = name
                break
        if not UI_FAMILY:
            UI_FAMILY = "Segoe UI"  # final fallback

    # Create shared fonts
    global FONT_MAIN, FONT_TITLE, FONT_BOLD
    FONT_MAIN = tkfont.Font(root, family=UI_FAMILY, size=10)
    FONT_TITLE = tkfont.Font(root, family=UI_FAMILY, size=11, weight="bold")
    FONT_BOLD = tkfont.Font(root, family=UI_FAMILY, size=10, weight="bold")

# ---------- Database ----------
def init_db():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("""
    CREATE TABLE IF NOT EXISTS history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        date TEXT
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS numbers (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        history_id INTEGER,
        number TEXT,
        FOREIGN KEY(history_id) REFERENCES history(id)
    )
    """)
    conn.commit()
    conn.close()

def save_to_history(numbers):
    if not numbers:
        return
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    when = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    c.execute("INSERT INTO history(date) VALUES(?)", (when,))
    history_id = c.lastrowid
    for num in numbers:
        c.execute("INSERT INTO numbers(history_id, number) VALUES(?, ?)", (history_id, num))
    conn.commit()
    conn.close()

def load_history():
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT id, date FROM history ORDER BY id ASC")
    rows = c.fetchall()
    conn.close()
    return rows

def load_numbers(history_id):
    conn = sqlite3.connect(DB_FILE)
    c = conn.cursor()
    c.execute("SELECT number FROM numbers WHERE history_id=?", (history_id,))
    nums = [r[0] for r in c.fetchall()]
    conn.close()
    return nums

# ---------- Helpers ----------
def normalize_number(raw):
    s = str(raw).translate(PERSIAN_TO_EN)
    s = re.sub(r'\D', '', s)
    if not s:
        return None
    if s.startswith('0'):
        s = '98' + s[1:]
    elif not s.startswith('98'):
        s = '98' + s
    return s

def update_output(numbers):
    global extracted_numbers
    numbers = [x for x in numbers if x]
    unique = sorted(set(numbers), key=lambda z: int(z))
    extracted_numbers = unique
    prefix = name_prefix_var.get().strip() or "contact"

    output_box.delete("1.0", tk.END)
    for i, num in enumerate(unique, 1):
        output_box.insert(tk.END, f"{prefix} - {i} +{num}\n")

    lbl_status.config(text=("📌 تعداد شماره‌ها: " if LANG_FA else "📌 Count: ") + str(len(unique)))
    if unique:
        save_to_history(unique)

# ---------- Inputs ----------
def extract_numbers_from_text():
    text = input_box.get("1.0", tk.END)
    found = [normalize_number(m.group(0)) for m in pattern.finditer(text)]
    update_output(found)

def extract_numbers_from_excel():
    filepath = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if not filepath:
        return
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        found = []
        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if cell:
                    nums = pattern.findall(str(cell))
                    for num in nums:
                        found.append(normalize_number(num))
        update_output(found)
        messagebox.showinfo("✅", tr("loaded_excel", filepath))
    except Exception as e:
        messagebox.showerror(tr("error"), str(e))

def extract_numbers_from_csv():
    filepath = filedialog.askopenfilename(filetypes=[("CSV files", "*.csv")])
    if not filepath:
        return
    try:
        found = []
        with open(filepath, newline="", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                for cell in row:
                    nums = pattern.findall(str(cell))
                    for num in nums:
                        found.append(normalize_number(num))
        update_output(found)
        messagebox.showinfo("✅", tr("loaded_csv", filepath))
    except Exception as e:
        messagebox.showerror(tr("error"), str(e))

# ---------- Outputs ----------
def save_to_file(filetype, numbers=None, custom_prefix=None):
    data = numbers if numbers is not None else extracted_numbers
    if not data:
        messagebox.showwarning("⚠️", tr("warn_no_numbers"))
        return

    prefix = (custom_prefix if custom_prefix is not None else name_prefix_var.get()).strip() or "contact"

    if filetype == "vcf":
        filepath = filedialog.asksaveasfilename(defaultextension=".vcf", filetypes=[("VCF files", "*.vcf")])
        if filepath:
            with open(filepath, "w", encoding="utf-8") as f:
                for i, num in enumerate(data, 1):
                    f.write("BEGIN:VCARD\r\nVERSION:3.0\r\n")
                    f.write(f"FN:{i}{prefix}\r\nTEL;TYPE=CELL:+{num}\r\nEND:VCARD\r\n")
            messagebox.showinfo("💾", tr("saved_vcf", filepath))

    elif filetype == "csv":
        filepath = filedialog.asksaveasfilename(defaultextension=".csv", filetypes=[("CSV files", "*.csv")])
        if filepath:
            with open(filepath, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(["Name", "Phone"])
                for i, num in enumerate(data, 1):
                    writer.writerow([f"{i}{prefix}", f"+{num}"])
            messagebox.showinfo("💾", tr("saved_csv", filepath))

    elif filetype == "xlsx":
        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
        if filepath:
            wb = Workbook()
            ws = wb.active
            ws.append(["Name", "Phone"])
            for i, num in enumerate(data, 1):
                ws.append([f"{i}{prefix}", f"+{num}"])
            wb.save(filepath)
            messagebox.showinfo("💾", tr("saved_xlsx", filepath))

def copy_to_clipboard(text=None):
    content = text if text is not None else output_box.get("1.0", tk.END).strip()
    if content:
        root.clipboard_clear()
        root.clipboard_append(content)
        messagebox.showinfo("📌", tr("copied"))

def paste_from_clipboard():
    try:
        text = root.clipboard_get()
        input_box.delete("1.0", tk.END)
        input_box.insert(tk.END, text)
    except tk.TclError:
        messagebox.showwarning("⚠️", tr("clipboard_empty"))

def clear_all():
    input_box.delete("1.0", tk.END)
    output_box.delete("1.0", tk.END)
    lbl_status.config(text=tr("status_ready"))

# ---------- History Window ----------
def open_history_window():
    hist_win = tk.Toplevel(root)
    hist_win.title(tr("history_window_title"))
    hist_win.configure(bg=BG_COLOR)
    hist_win.geometry("900x500")

    # ttk style (dark)
    style = ttk.Style(hist_win)
    style.theme_use("default")
    style.configure("Treeview",
                    background=TEXT_BG, foreground=TEXT_FG,
                    fieldbackground=TEXT_BG, rowheight=24, font=FONT_MAIN)
    style.configure("Treeview.Heading",
                    background=BTN_COLOR, foreground="white", font=FONT_TITLE)
    style.map("Treeview", background=[("selected", BTN_HOVER)])

    # Split view
    left = tk.Frame(hist_win, bg=BG_COLOR)
    left.pack(side="left", fill="y", padx=8, pady=8)

    right = tk.Frame(hist_win, bg=BG_COLOR)
    right.pack(side="left", fill="both", expand=True, padx=8, pady=8)

    # Left: dates tree
    tk.Label(left, text=tr("history_dates"), bg=BG_COLOR, fg="white", font=FONT_TITLE).pack(anchor="w", pady=(0,6))
    tree = ttk.Treeview(left, columns=("id", "date"), show="headings", height=18)
    tree.heading("id", text=tr("history_id"))
    tree.heading("date", text=tr("history_datetime"))
    tree.column("id", width=60, anchor="center")
    tree.column("date", width=200)
    tree.pack(fill="y")

    # Fill dates
    for rec in load_history():
        tree.insert("", "end", values=rec)

    # Right: numbers + toolbar
    top_right = tk.Frame(right, bg=BG_COLOR)
    top_right.pack(fill="x")

    tk.Label(top_right, text=tr("history_prefix"), bg=BG_COLOR, fg="white", font=FONT_MAIN).pack(side="left")
    hist_prefix_var = tk.StringVar(value=(name_prefix_var.get() or "contact"))
    tk.Entry(top_right, textvariable=hist_prefix_var, width=18, font=FONT_MAIN).pack(side="left", padx=6)

    count_lbl = tk.Label(top_right, text="—", bg=BG_COLOR, fg="lightgray", font=FONT_MAIN)
    count_lbl.pack(side="right")

    nums_box = scrolledtext.ScrolledText(right, width=60, height=20, bg=TEXT_BG, fg=TEXT_FG, font=FONT_MAIN, insertbackground="white")
    nums_box.pack(fill="both", expand=True, pady=6)

    toolbar = tk.Frame(right, bg=BG_COLOR)
    toolbar.pack(fill="x")

    selected_session_numbers = {"list": []}

    def build_preview_text(nums):
        prefix = (hist_prefix_var.get().strip() or "contact")
        lines = [f"{prefix} - {i} +{n}" for i, n in enumerate(nums, 1)]
        return "\n".join(lines)

    def on_select(_event=None):
        sel = tree.focus()
        if not sel:
            return
        values = tree.item(sel, "values")
        if not values:
            return
        hist_id = int(values[0])
        nums = load_numbers(hist_id)
        selected_session_numbers["list"] = nums

        nums_box.config(state="normal")
        nums_box.delete("1.0", tk.END)
        nums_box.insert(tk.END, build_preview_text(nums))
        nums_box.config(state="normal")

        count_lbl.config(text=tr("history_count", len(nums)))

    def copy_session():
        if not selected_session_numbers["list"]:
            messagebox.showwarning("⚠️", tr("history_select_first"))
            return
        copy_to_clipboard(build_preview_text(selected_session_numbers["list"]))

    def export_session(kind):
        if not selected_session_numbers["list"]:
            messagebox.showwarning("⚠️", tr("history_select_first"))
            return
        save_to_file(kind, numbers=selected_session_numbers["list"], custom_prefix=hist_prefix_var.get())

    def make_btn(parent, text, cmd):
        b = tk.Button(parent, text=text, command=cmd, bg=BTN_COLOR, fg="white",
                      relief="flat", padx=10, pady=6, font=FONT_MAIN)
        b.pack(side="left", padx=5, pady=2)
        b.bind("<Enter>", lambda e: b.config(bg=BTN_HOVER))
        b.bind("<Leave>", lambda e: b.config(bg=BTN_COLOR))
        return b

    make_btn(toolbar, tr("btn_copy"), copy_session)
    make_btn(toolbar, tr("btn_save_vcf"), lambda: export_session("vcf"))
    make_btn(toolbar, tr("btn_save_csv"), lambda: export_session("csv"))
    make_btn(toolbar, tr("btn_save_xlsx"), lambda: export_session("xlsx"))

    tree.bind("<<TreeviewSelect>>", on_select)

    def refresh_preview_on_prefix_change(*_):
        if selected_session_numbers["list"]:
            nums_box.config(state="normal")
            nums_box.delete("1.0", tk.END)
            nums_box.insert(tk.END, build_preview_text(selected_session_numbers["list"]))
            nums_box.config(state="normal")
    hist_prefix_var.trace_add("write", refresh_preview_on_prefix_change)

# ---------- UI ----------
root = tk.Tk()
load_ui_font(root)

# Apply global options for fonts/colors
root.option_add("*Font", FONT_MAIN)
root.option_add("*Label.Font", FONT_MAIN)
root.option_add("*Button.Font", FONT_MAIN)
root.option_add("*Entry.Font", FONT_MAIN)
root.option_add("*TCombobox*Listbox.Font", FONT_MAIN)

root.title(T["title"][1] if LANG_FA else T["title"][0])
root.configure(bg=BG_COLOR)

# Menu bar
menubar = tk.Menu(root, tearoff=0, font=FONT_MAIN)
file_menu = tk.Menu(menubar, tearoff=0, font=FONT_MAIN)
file_menu.add_command(label=tr("menu_exit"), command=root.quit)
menubar.add_cascade(label=tr("menu_file"), menu=file_menu)

history_menu = tk.Menu(menubar, tearoff=0, font=FONT_MAIN)
history_menu.add_command(label=tr("menu_history_show"), command=open_history_window)
menubar.add_cascade(label=tr("menu_history"), menu=history_menu)

lang_menu = tk.Menu(menubar, tearoff=0, font=FONT_MAIN)
lang_menu.add_command(label=tr("menu_lang_toggle"), command=toggle_language)
menubar.add_cascade(label=tr("menu_lang"), menu=lang_menu)
root.config(menu=menubar)

# Input text
lbl_input = tk.Label(root, text=tr("label_input"), bg=BG_COLOR, fg="white", font=FONT_TITLE)
lbl_input.pack(anchor="w", padx=8, pady=(8,3))
input_box = scrolledtext.ScrolledText(root, width=86, height=6, bg=TEXT_BG, fg=TEXT_FG, font=FONT_MAIN, insertbackground="white", relief="flat", bd=0)
input_box.pack(padx=10, pady=5, fill="x")

# Name prefix
prefix_frame = tk.Frame(root, bg=BG_COLOR)
prefix_frame.pack(anchor="w", padx=10, pady=(0,6))
lbl_prefix = tk.Label(prefix_frame, text=tr("label_prefix"), bg=BG_COLOR, fg="white", font=FONT_MAIN)
lbl_prefix.pack(side="left")
name_prefix_var = tk.StringVar(value="contact")
tk.Entry(prefix_frame, textvariable=name_prefix_var, width=18, font=FONT_MAIN, relief="flat").pack(side="left", padx=6)

# Panels
input_frame = tk.LabelFrame(root, text=tr("group_input"), bg=BG_COLOR, fg="white", font=FONT_TITLE, labelanchor="n")
input_frame.pack(fill="x", padx=10, pady=5)
output_frame = tk.LabelFrame(root, text=tr("group_output"), bg=BG_COLOR, fg="white", font=FONT_TITLE, labelanchor="n")
output_frame.pack(fill="x", padx=10, pady=5)

def make_button(parent, text, cmd):
    btn = tk.Button(parent, text=text, command=cmd, bg=BTN_COLOR, fg="white", font=FONT_MAIN, relief="flat", padx=10, pady=6)
    btn.pack(side="left", padx=6, pady=6)
    btn.bind("<Enter>", lambda e: btn.config(bg=BTN_HOVER))
    btn.bind("<Leave>", lambda e: btn.config(bg=BTN_COLOR))
    return btn

# ---------- Input panel buttons ----------
btn_paste = make_button(input_frame, tr("btn_paste"), paste_from_clipboard)
btn_load_excel = make_button(input_frame, tr("btn_load_excel"), extract_numbers_from_excel)
btn_load_csv = make_button(input_frame, tr("btn_load_csv"), extract_numbers_from_csv)

# ---------- Output panel buttons ----------
btn_extract_text = make_button(output_frame, tr("btn_extract_text"), extract_numbers_from_text)
btn_save_vcf = make_button(output_frame, tr("btn_save_vcf"), lambda: save_to_file("vcf"))
btn_save_csv = make_button(output_frame, tr("btn_save_csv"), lambda: save_to_file("csv"))
btn_save_xlsx = make_button(output_frame, tr("btn_save_xlsx"), lambda: save_to_file("xlsx"))
btn_copy = make_button(output_frame, tr("btn_copy"), lambda: copy_to_clipboard())
btn_history = make_button(output_frame, tr("btn_history"), open_history_window)
btn_lang_inline = make_button(output_frame, tr("btn_lang"), toggle_language)

# ---------- Separate Utility Buttons ----------
util_frame = tk.Frame(output_frame, bg=BG_COLOR)
util_frame.pack(fill="x", pady=6)

def open_donate_window():
    donate_win = tk.Toplevel(root)
    donate_win.title(tr("donate_title"))
    donate_win.configure(bg=BG_COLOR)
    donate_win.geometry("800x800")
    donate_win.resizable(False, False)

    tk.Label(donate_win, text=tr("donate_thanks"), bg=BG_COLOR, fg="white", font=FONT_TITLE).pack(pady=(10,5))
    tk.Label(donate_win, text=tr("donate_hint"), bg=BG_COLOR, fg="lightgray", font=FONT_MAIN).pack(pady=(0,10))

    donations = [
        ("Tether (USDT)", "TFGov47L6aK8MbtuiPesgvZBj4VTARKrBQ", False),
        ("Tron (TRX)", "TFGov47L6aK8MbtuiPesgvZBj4VTARKrBQ", False),
        ("Bitcoin (BTC)", "bc1q20hd42vkefpwt75h3pp3uj2nsfu0xvtpu2dgrl", False),
        ("Coffeete", "https://www.coffeete.ir/RezaZare", True)
    ]

    frame_grid = tk.Frame(donate_win, bg=BG_COLOR)
    frame_grid.pack(pady=10)

    for i, (name, address, is_link) in enumerate(donations):
        row = i // 2
        col = i % 2
        frame = tk.Frame(frame_grid, bg=BG_COLOR, padx=10, pady=10, relief="ridge", bd=1)
        frame.grid(row=row, column=col, padx=10, pady=10)

        tk.Label(frame, text=name, bg=BG_COLOR, fg="white", font=FONT_TITLE).pack()

        addr_label = tk.Label(frame, text=address, bg=BG_COLOR, fg="#7CFC00", font=FONT_MAIN)
        addr_label.pack(pady=(2,5))

        def make_copy(addr=address, name=name):
            donate_win.clipboard_clear()
            donate_win.clipboard_append(addr)
            messagebox.showinfo("📌", tr("copied"))

        tk.Button(frame, text=tr("btn_copy"), command=make_copy,
                  bg=BTN_COLOR, fg="white", font=FONT_MAIN, relief="flat", padx=6, pady=3).pack()

        if is_link:
            def open_web(event, link=address):
                webbrowser.open(link)
            link_label = tk.Label(frame, text=tr("link_open"), bg=BG_COLOR, fg="#FFD700",
                                  cursor="hand2", font=FONT_MAIN)
            link_label.pack(pady=(2,5))
            link_label.bind("<Button-1>", open_web)

        qr = qrcode.QRCode(version=1, box_size=4, border=1)
        qr.add_data(address)
        qr.make(fit=True)
        img = qr.make_image(fill_color="black", back_color="white")
        img = img.resize((150,150))
        qr_img = ImageTk.PhotoImage(img)
        qr_label = tk.Label(frame, image=qr_img, bg=BG_COLOR)
        qr_label.image = qr_img
        qr_label.pack(pady=5)
    
    tk.Button(donate_win, text=tr("close"), command=donate_win.destroy, 
              bg=BTN_COLOR, fg="white", font=FONT_TITLE, relief="flat", padx=10, pady=5).pack(pady=10)

make_button(util_frame, tr("btn_clear"), clear_all)
make_button(util_frame, tr("btn_donate"), open_donate_window)

# Output preview
lbl_preview = tk.Label(root, text=tr("label_preview"), bg=BG_COLOR, fg="white", font=FONT_TITLE)
lbl_preview.pack(anchor="w", padx=8, pady=(6,3))
output_box = scrolledtext.ScrolledText(root, width=86, height=12, bg=TEXT_BG, fg=TEXT_FG, font=FONT_MAIN, insertbackground="white")
output_box.pack(padx=10, pady=5, fill="both", expand=True)

lbl_status = tk.Label(root, text=tr("status_ready"), bg=BG_COLOR, fg="lightgray", font=FONT_MAIN)
lbl_status.pack(pady=6)

extracted_numbers = []

def refresh_ui_texts():
    root.title(tr("title"))
    # Menus
    menubar.entryconfig(0, label=tr("menu_file"))
    menubar.entryconfig(1, label=tr("menu_history"))
    menubar.entryconfig(2, label=tr("menu_lang"))
    file_menu.entryconfig(0, label=tr("menu_exit"))
    history_menu.entryconfig(0, label=tr("menu_history_show"))
    lang_menu.entryconfig(0, label=tr("menu_lang_toggle"))
    # Labels / frames
    lbl_input.config(text=tr("label_input"), font=FONT_TITLE)
    lbl_prefix.config(text=tr("label_prefix"))
    input_frame.config(text=tr("group_input"), font=FONT_TITLE)
    output_frame.config(text=tr("group_output"), font=FONT_TITLE)
    lbl_preview.config(text=tr("label_preview"))
    lbl_status.config(text=tr("status_ready"))
    # Buttons
    btn_paste.config(text=tr("btn_paste"))
    btn_extract_text.config(text=tr("btn_extract_text"))
    btn_load_excel.config(text=tr("btn_load_excel"))
    btn_load_csv.config(text=tr("btn_load_csv"))
    btn_save_vcf.config(text=tr("btn_save_vcf"))
    btn_save_csv.config(text=tr("btn_save_csv"))
    btn_save_xlsx.config(text=tr("btn_save_xlsx"))
    btn_copy.config(text=tr("btn_copy"))
    btn_history.config(text=tr("btn_history"))
    btn_lang_inline.config(text=tr("btn_lang"))

# Build menus after function reference exists
menubar = tk.Menu(root, tearoff=0, font=FONT_MAIN)
file_menu = tk.Menu(menubar, tearoff=0, font=FONT_MAIN)
file_menu.add_command(label=tr("menu_exit"), command=root.quit)
menubar.add_cascade(label=tr("menu_file"), menu=file_menu)

history_menu = tk.Menu(menubar, tearoff=0, font=FONT_MAIN)
history_menu.add_command(label=tr("menu_history_show"), command=open_history_window)
menubar.add_cascade(label=tr("menu_history"), menu=history_menu)

lang_menu = tk.Menu(menubar, tearoff=0, font=FONT_MAIN)
lang_menu.add_command(label=tr("menu_lang_toggle"), command=toggle_language)
menubar.add_cascade(label=tr("menu_lang"), menu=lang_menu)
root.config(menu=menubar)

init_db()
refresh_ui_texts()
root.mainloop()
